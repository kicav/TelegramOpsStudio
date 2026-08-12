from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

FIELDS = [
    "user_id", "access_hash", "username", "first_name", "last_name", "phone",
    "is_bot", "is_deleted", "has_photo", "last_seen", "source_group", "source_managed",
    "consent_status", "consent_note", "status", "last_error",
]


def _value(row, key: str):
    if hasattr(row, "keys"):
        return row[key] if key in row.keys() else ""
    return row.get(key, "")


def export_csv(rows, path: str) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _value(row, key) for key in FIELDS})


def export_xlsx(rows, path: str) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(FIELDS)
    for row in rows:
        ws.append([_value(row, key) for key in FIELDS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        width = max(10, min(42, max(len(str(cell.value or "")) for cell in column) + 2))
        ws.column_dimensions[column[0].column_letter].width = width
    wb.save(target)


def import_xlsx(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration:
            return []
        headers = [str(value or "").strip() for value in first]
        if "user_id" not in headers:
            raise ValueError("XLSX is missing required 'user_id' column")
        result: list[dict] = []
        for values in rows:
            row = dict(zip(headers, values))
            if row.get("user_id") not in (None, ""):
                result.append(row)
        return result
    finally:
        wb.close()
